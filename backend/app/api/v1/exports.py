"""Bulk export endpoints — CSV and Excel download for grades and students."""
import csv
import io
from datetime import datetime

from flask import Blueprint, Response, jsonify, request, g, send_file
from flask_jwt_extended import get_jwt_identity, jwt_required
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from ...core.database import session_scope
from ...core.decorators import require_roles
from ...core.extensions import limiter
from ...models import Aluno, Nota, Tenant
from ...services.access_notice_service import AccessNoticeService
from ...services.audit import log_action


def _safe(value) -> str:
    """Convert any value to a safe string for CSV/Excel cells."""
    if value is None:
        return ""
    text = str(value)
    # Prevent spreadsheet formula injection when CSV/XLSX is opened by staff.
    if text.startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + text
    return text


def _float_or_empty(value) -> str:
    if value is None:
        return ""
    try:
        return f"{float(value):.1f}"
    except (TypeError, ValueError):
        return ""


def _safe_filename(prefix: str) -> str:
    return secure_filename(prefix) or "export"


def _apply_download_headers(response: Response) -> Response:
    response.headers["Cache-Control"] = "no-store, no-cache, private"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


def register(parent: Blueprint) -> None:
    bp = Blueprint("exports", __name__)

    # ── Students export ──────────────────────────────────────────────────────

    @bp.get("/exports/alunos")
    @jwt_required()
    @limiter.limit("20 per hour")
    @require_roles("admin", "super_admin", "coordenador", "diretor", "orientador", "professor")
    def export_alunos():
        """Download the full student list with grade averages.

        Query params:
            format  csv | xlsx   (default: csv)
            turma   filter by class name
            turno   filter by shift
        """
        fmt = (request.args.get("format") or "csv").lower()
        turma_filter = (request.args.get("turma") or "").strip() or None
        turno_filter = (request.args.get("turno") or "").strip() or None
        user_id = int(get_jwt_identity())

        if fmt not in ("csv", "xlsx"):
            return jsonify({"error": "Formato inválido. Use 'csv' ou 'xlsx'."}), 400

        with session_scope() as session:
            query = (
                session.query(Aluno)
                .options(joinedload(Aluno.notas))
                .filter(Aluno.tenant_id == g.tenant_id)
            )
            if getattr(g, "academic_year_id", None):
                query = query.filter(Aluno.academic_year_id == g.academic_year_id)
            if turma_filter:
                query = query.filter(Aluno.turma == turma_filter)
            if turno_filter:
                query = query.filter(Aluno.turno == turno_filter)

            alunos = query.order_by(Aluno.turma, Aluno.nome).all()

            rows = []
            for a in alunos:
                notas_vals = [float(n.total) for n in a.notas if n.total is not None]
                media = round(sum(notas_vals) / len(notas_vals), 1) if notas_vals else None
                total_faltas = sum(n.faltas or 0 for n in a.notas)
                rows.append({
                    "Matrícula": _safe(a.matricula),
                    "Nome": _safe(a.nome),
                    "Turma": _safe(a.turma),
                    "Turno": _safe(a.turno),
                    "Status": _safe(a.status) or "Ativo",
                    "Média Geral": _float_or_empty(media),
                    "Total de Faltas": _safe(total_faltas),
                })

            log_action(
                session,
                user_id,
                "EXPORT_ALUNOS",
                "Aluno",
                details={
                    "format": fmt,
                    "row_count": len(rows),
                    "academic_year_id": getattr(g, "academic_year_id", None),
                    "filters": {"turma": turma_filter, "turno": turno_filter},
                },
            )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = _safe_filename(f"alunos_{timestamp}")

        if fmt == "csv":
            return _csv_response(rows, filename)
        return _xlsx_response(rows, "Alunos", filename)

    # ── Grades export ─────────────────────────────────────────────────────────

    @bp.get("/exports/notas")
    @jwt_required()
    @limiter.limit("20 per hour")
    @require_roles("admin", "super_admin", "coordenador", "diretor", "orientador", "professor")
    def export_notas():
        """Download the complete grades table.

        Query params:
            format      csv | xlsx          (default: csv)
            turma       filter by class
            turno       filter by shift
            disciplina  filter by subject
        """
        fmt = (request.args.get("format") or "csv").lower()
        turma_filter = (request.args.get("turma") or "").strip() or None
        turno_filter = (request.args.get("turno") or "").strip() or None
        disciplina_filter = (request.args.get("disciplina") or "").strip() or None
        user_id = int(get_jwt_identity())

        if fmt not in ("csv", "xlsx"):
            return jsonify({"error": "Formato inválido. Use 'csv' ou 'xlsx'."}), 400

        with session_scope() as session:
            query = (
                session.query(Nota)
                .options(joinedload(Nota.aluno))
                .filter(Nota.tenant_id == g.tenant_id)
            )
            if getattr(g, "academic_year_id", None):
                query = query.filter(Nota.academic_year_id == g.academic_year_id)
            if disciplina_filter:
                query = query.filter(Nota.disciplina == disciplina_filter)
            if turma_filter or turno_filter:
                query = query.join(Aluno)
                if turma_filter:
                    query = query.filter(Aluno.turma == turma_filter)
                if turno_filter:
                    query = query.filter(Aluno.turno == turno_filter)

            notas = (
                query.order_by(Aluno.turma, Aluno.nome, Nota.disciplina).all()
                if (turma_filter or turno_filter)
                else query.order_by(Nota.disciplina).all()
            )

            rows = []
            for n in notas:
                aluno = n.aluno
                rows.append({
                    "Matrícula": _safe(aluno.matricula if aluno else ""),
                    "Aluno": _safe(aluno.nome if aluno else ""),
                    "Turma": _safe(aluno.turma if aluno else ""),
                    "Turno": _safe(aluno.turno if aluno else ""),
                    "Disciplina": _safe(n.disciplina),
                    "1º Trimestre": _float_or_empty(n.trimestre1),
                    "2º Trimestre": _float_or_empty(n.trimestre2),
                    "3º Trimestre": _float_or_empty(n.trimestre3),
                    "Média Final": _float_or_empty(n.total),
                    "Faltas": _safe(n.faltas),
                    "Situação": _safe(n.situacao),
                })

            log_action(
                session,
                user_id,
                "EXPORT_NOTAS",
                "Nota",
                details={
                    "format": fmt,
                    "row_count": len(rows),
                    "academic_year_id": getattr(g, "academic_year_id", None),
                    "filters": {
                        "turma": turma_filter,
                        "turno": turno_filter,
                        "disciplina": disciplina_filter,
                    },
                },
            )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = _safe_filename(f"notas_{timestamp}")

        if fmt == "csv":
            return _csv_response(rows, filename)
        return _xlsx_response(rows, "Notas", filename)

    # ── Access notices ────────────────────────────────────────────────────────

    @bp.get("/exports/comunicados-acesso")
    @jwt_required()
    @require_roles(
        "admin",
        "super_admin",
        "coordenador",
        "coordenacao",
        "diretor",
        "orientador",
        "orientacao",
    )
    def export_comunicados_acesso():
        """Download guardian or student access notices as one DOCX per class."""
        turma_filter = (request.args.get("turma") or "").strip()
        if not turma_filter:
            return jsonify({"error": "Informe a turma para gerar os comunicados."}), 400

        tipo = (request.args.get("tipo") or "responsavel").strip().lower()
        if tipo not in ("responsavel", "aluno"):
            tipo = "responsavel"

        tenant_id = getattr(g, "tenant_id", None)
        academic_year_id = getattr(g, "academic_year_id", None)
        if not tenant_id:
            return jsonify({"error": "Escola não identificada."}), 400

        user_id = int(get_jwt_identity())
        with session_scope() as session:
            tenant = session.get(Tenant, tenant_id)
            school_name = tenant.name if tenant else "ColaboraEdu"
            try:
                result = AccessNoticeService(session).generate_class_docx(
                    tenant_id=tenant_id,
                    academic_year_id=academic_year_id,
                    turma=turma_filter,
                    school_name=school_name,
                    tipo=tipo,
                )
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 404
            except RuntimeError as exc:
                return jsonify({"error": str(exc)}), 500

            log_action(
                session,
                user_id,
                "EXPORT_COMUNICADOS_ACESSO",
                "Aluno",
                details={
                    "turma": turma_filter,
                    "row_count": result.row_count,
                    "academic_year_id": academic_year_id,
                    "format": "docx",
                    "passwords_reset": True,
                    "tipo": tipo,
                },
            )

        response = send_file(
            result.buffer,
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            as_attachment=True,
            download_name=secure_filename(result.filename) or "comunicados_acesso.docx",
        )
        return _apply_download_headers(response)

    parent.register_blueprint(bp)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _csv_response(rows: list[dict], filename: str) -> Response:
    if not rows:
        rows = [{}]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()), lineterminator="\r\n")
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    response = Response(
        output.getvalue().encode("utf-8-sig"),  # BOM so Excel opens correctly
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )
    return _apply_download_headers(response)


def _xlsx_response(rows: list[dict], sheet_name: str, filename: str) -> Response:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel tab name limit

    if not rows:
        rows = [{}]

    headers = list(rows[0].keys())

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_safe(row.get(header, "")))

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(header),
            *(len(str(r.get(header, "") or "")) for r in rows),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )
    return _apply_download_headers(response)
